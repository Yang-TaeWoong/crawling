from openpyxl import Workbook
from bs4 import BeautifulSoup
from selenium import webdriver
from datetime import date


options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
options.add_argument("lang=ko_KR")    # 가짜 플러그인 탑재
options.add_argument("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/71.0.3578.98 Safari/537.36")

# driver = webdriver.Chrome('C:\python/chromedriver',options=options)
checkList_info=("아이디","이름","집주소","전화번호","통관번호")
driver = webdriver.Chrome('C:\python/chromedriver')
customer_info_arr=[] #Customer 클래스의 배열

class Customer:

    def __init__(self,id,name,address,phone,custom_num):
        self.id=id
        self.name=name
        self.address=address
        self.phone=phone
        self.custom_num=custom_num
    def get_customer_info(self):
        info=[0,self.id,self.name,self.address,self.phone,self.custom_num]
        return info

    def set_cutomer_info(self,id,name,address,phone,custom_num):
        self.id=id
        self.name=name
        self.address=address
        self.phone=phone
        self.custom_num=custom_num

    # options = webdriver.ChromeOptions()
    # options.add_argument('headless')
    # options.add_argument('window-size=1920x1080')
    # options.add_argument("disable-gpu")
    # driver = webdriver.Chrome('C:\python/chromedriver', options=options)
    # 헤드리스 모드 진행시 블록아웃을 삭제

def make_excel(url, customer_count):
    wb = Workbook()
    ws1 = wb.active
    # ws2= wb.create_sheet(title="second_sheet")  새로운 시트 파일 하나 더 추가하기
    for col in range(1,6):
        ws1.cell(row=1,column=col,value=checkList_info[col-1])
    # 이름, 전번, 집주소, 통관번호 입력을 넣어준다
    for row in range(2, customer_count + 2):
        customer_info=customer_info_arr[row - 2].get_customer_info()
        for col in range(1, 6):
            ws1.cell(row=row, column=col).value=customer_info[col]  # TODO : value값에 그냥 숫자들이 아니라 개인 신상들어가야함
    wb.save(url)
    wb.close()

def get_excel_file_name():
     return date.today().isoformat()+'.xlsx'


def add_customer(id,name,address,phone_num,custom_num):
    customer=Customer(id,name,address,phone_num,custom_num)
    customer_info_arr.append(customer)

def open_info_pages(url):
    driver.implicitly_wait(3)
    driver.get(url)
    read_spec_order()
    #     TODO 여기 부분 수정해야한다. 링크를 클릭하면 창들이 탭으로 뜨면서 정보들을 크롤링해오는 방향으로 간다

def read_spec_order():

    id=driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[1]/tbody/tr[4]/td').text
    name=driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[1]/td').text
    phone_num = driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[2]/td[1]').text
    address=driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[3]/td').text
    custom_num=driver.find_element_by_xpath('//*[@id="pop_content"]/div/table[3]/tbody/tr[4]/td').text
    print(id,name,address,phone_num,custom_num)
    add_customer(id,name, address, phone_num, custom_num)

def read_order_and_return_customer_count():
    customer_count = 0
    # url=driver.current_url  #이부분에서 현재 url을 긁어온다
    # TODO: 여기서 리스트를 넣어주지 말고 읽어와서 처리하도록
    url= 'https://sell.smartstore.naver.com/o/orderDetail/productOrder/' + '2019011916399251' + '/popup' #상세페이지 url
    open_info_pages(url)
    customer_count+=1
    # for i in customer_count:
    return customer_count

def log_in():
    try:
        driver.implicitly_wait(3)
        driver.get('https://sell.smartstore.naver.com/#/login')
        driver.find_element_by_id('loginId').send_keys('ytw1122@gmail.com')
        driver.find_element_by_id('loginPassword').send_keys('andrew3876')
        driver.find_element_by_xpath('//*[@id="loginButton"]').click()
    except Exception as inst:
        print("fuck, its not working")
        print(inst.args)
#     in case when login doesnt work properly

def go_to_order_page():
    drop_down_menu = driver.find_element_by_xpath('//*[@id="seller-lnb"]/div/div[1]/ul/li[3]')
    drop_down_menu.click()
    driver.implicitly_wait(3)
    driver.find_element_by_css_selector('#seller-lnb > div > div:nth-child(1) > ul > li.active > ul > li:nth-child(1) > a').click()

log_in()
driver.implicitly_wait(3)
go_to_order_page()
customer_count=read_order_and_return_customer_count()
file_name="C:\study/"+get_excel_file_name()
make_excel(file_name,customer_count)
# 3의 의미 : 데이터의 갯수가 3개다
